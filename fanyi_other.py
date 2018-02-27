from openpyxl import Workbook, load_workbook
import hashlib
import requests
import re
from collections import deque
import os, time,sys
import trip



def get_fanyi(q):
    pattern = re.compile(r'[\u4e00-\u9fa5]')
    salt = '123'
    appid = '20180115000115453'
    secret = 'AHYyafcYLzm3q0fkfQ7z'
    s = appid + q + salt + secret
    sign = string_to_md5(s)
    if pattern.findall(q):
        url = 'http://api.fanyi.baidu.com/api/trans/vip/translate?' \
              'q={}&from=zh&to=en&appid={}&salt={}&sign={}'.format(q, appid, salt, sign)
    else:
        url = 'http://api.fanyi.baidu.com/api/trans/vip/translate?' \
              'q={}&from=en&to=zh&appid={}&salt={}&sign={}'.format(q, appid, salt, sign)
    try:
        r = requests.get(url,timeout=60)
    except Exception as why:
        print(why)
        return


def string_to_md5(s):
    md5 = hashlib.md5(s.encode('utf-8')).hexdigest()
    return md5

def get_xlsx_name():
    files = os.listdir()
    for file in files:
        if file.split('.')[1] == 'xlsx':
            return file

def get_content(name):
    wb = load_workbook(str(name))
    ws1 = wb.active
    contents = []
    for i in ws1['A']:
        if i.value:
            contents.append(i.value)
    # print(contents)
    return contents

src = []
dst = []

def save_result():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'translate'
    for s, d in zip(src, dst):
        name_index = src.index(s) + 1
        ws1['A%d' % name_index].value = s
        ws1['B%d' % name_index].value = d
    wb.save('results.xlsx')

def main():
    file_name = get_xlsx_name()
    contents = get_content(file_name)
    n = 1
    try:
        for q in contents:
            result = get_fanyi(q)
            time.sleep(0.2)
            if result is None:
                continue
            trans_result= result.json()['trans_result'][0]
            print('正在翻译第{}条...'.format(n))
            src.append(trans_result['src'])
            dst.append(trans_result['dst'])
            n += 1
    except Exception as why:
        print(why)
        save_result()
    else:
        print('翻译完毕，正在写入Excel...')
        save_result()

if __name__ == '__main__':
    start = time.time()
    main()
    print('写入完成,共耗时{}'.format(time.time() - start))
    print('程序正在退出...')
    time.sleep(0.6)