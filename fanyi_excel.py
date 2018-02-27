import hashlib
import os
import re
import time

import grequests
from openpyxl import Workbook, load_workbook

#生成待翻译的url链接
def get_fanyi(q):
    pattern = re.compile(r'[\u4e00-\u9fa5]')
    salt = '1122'
    appid = '20180115000115453'
    secret = 'AHYyafcYLzm3q0fkfQ7z'
    s = appid + q + salt + secret
    sign = string_to_md5(s)
    if pattern.findall(q):
        url = 'http://api.fanyi.baidu.com/api/trans/vip/translate?' \
              'q={}&from=zh&to=en&appid={}&salt={}&sign={}'.format(
                  q, appid, salt, sign)
    else:
        url = 'http://api.fanyi.baidu.com/api/trans/vip/translate?' \
              'q={}&from=en&to=zh&appid={}&salt={}&sign={}'.format(
                  q, appid, salt, sign)

    return url
    # r =  requests.get(url, timeout=60)
    # return r

#字符串转MD5
def string_to_md5(s):
    md5 = hashlib.md5(s.encode('utf-8')).hexdigest()
    return md5

#从excel得到待翻译内容
def get_content(name):
    wb = load_workbook(str(name))
    ws1 = wb.active
    contents = []
    for i in ws1['A']:
        if i.value:
            contents.append(i.value)
    print(len(contents))
    return contents

#将关键词列表分段，一段一段翻译
def list_of_groups(init_list, childern_list_len):
    list_of_groups = zip(*(iter(init_list),) * childern_list_len)
    end_list = [list(i) for i in list_of_groups]
    count = len(init_list) % childern_list_len
    end_list.append(init_list[-count:]) if count != 0 else end_list
    return end_list

#获得当前文件夹待翻译xlsx文件
def get_xlsx_name():
    files = os.listdir('C:\\Users\\Administrator\\Desktop\\fanyi_api')
    file = [file for file in files if file.endswith('.xlsx')][0]
    return file


src = []
dst = []

#将翻译后的结果保存
def save_result():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'translate'
    for s, d in zip(src, dst):
        name_index = src.index(s) + 1
        ws1['A%d' % name_index].value = s
        ws1['B%d' % name_index].value = d
    wb.save('results.xlsx')


#
# def save_contents(contents):
#     for i in contents:
#         name_index = src.index(i) + 1

#运行主逻辑
def main():
    file_name = get_xlsx_name()
    contents = get_content(file_name)
    n = 1
    try:
        for each in list_of_groups(contents, 30):
            request_urls = [get_fanyi(q) for q in each]
            rs = (grequests.get(u) for u in request_urls)
            r = grequests.map(rs)
            print(r)
            trans_results = filter(lambda x: x, r)
            for result in trans_results:
                trans_result = result.json()['trans_result'][0]
                print('正在翻译第{}条...'.format(n))
                src.append(trans_result['src'])
                dst.append(trans_result['dst'])
                n += 1
    except Exception as why:
        print(why)
        print('出错了，已把翻译过的保存...')
        save_result()
    else:
        print('翻译完毕，正在写入Excel...')
        save_result()


if __name__ == '__main__':
    start = time.time()
    main()
    print('写入完成,共耗时{}'.format(time.time() - start))
    print('程序正在退出...')
    time.sleep(0.6)
